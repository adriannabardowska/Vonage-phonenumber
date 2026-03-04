import requests
import pandas as pd
import time
import base64
from datetime import datetime, timezone
import io
import warnings
import os

# Suppress warnings
warnings.filterwarnings("ignore")
pd.set_option('display.max_rows', None)
pd.set_option('display.width', 1000)
pd.set_option('display.colheader_justify', 'left')

# --- CONFIGURATION ---
API_KEY = os.getenv("VONAGE_API_KEY", "").strip()
API_SECRET = os.getenv("VONAGE_API_SECRET", "").strip()
START_DATE_HISTORY = os.getenv("START_DATE_HISTORY", "2025-11-01").strip()
START_DATE_SMS = os.getenv("START_DATE_SMS", "2025-11-01T00:00:00Z").strip()

# Price per country (EUR/month)
COUNTRY_PRICES = {'US': 0.93, 'CA': 0.72, 'GB': 1.00}

# Slack Webhook URL for notifications
SLACK_WEBHOOK_URL = os.getenv("SLACK_WEBHOOK_URL", "").strip()

# Snapshot file for stable historical data
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SNAPSHOT_FILE = os.path.join(SCRIPT_DIR, "vonage_numbers_snapshot.csv")

# --- ENDPOINTS ---
BASE_URL_REST = "https://rest.nexmo.com"
BASE_URL_API = "https://api.nexmo.com"


def validate_config():
    missing = []
    if not API_KEY:
        missing.append("VONAGE_API_KEY")
    if not API_SECRET:
        missing.append("VONAGE_API_SECRET")
    if missing:
        raise RuntimeError(
            "Missing required environment variables: " + ", ".join(missing)
        )


def send_slack_notification(breakdown_df, utilization_df, excel_filename):
    """Send a summary report to Slack."""
    if not SLACK_WEBHOOK_URL:
        print("   Slack webhook not configured, skipping notification.")
        return False

    try:
        # Get latest month data for summary
        total_rows = breakdown_df[breakdown_df['Country'] == '>>> TOTAL'].copy()

        # Build summary table
        summary_lines = []
        for _, row in total_rows.iterrows():
            new_activations = int(row.get('New Activations', 0))
            net_change = int(row.get('Net Change', 0))
            cancelled_count = int(row.get('Cancelled This Month', new_activations - net_change))
            summary_lines.append(
                f"*{row['Month']}*: {row['Active Numbers']} active | "
                f"+{new_activations} new | -{cancelled_count} cancelled | "
                f"Net: {net_change}"
            )

        # Get utilization summary
        util_lines = []
        if not utilization_df.empty:
            for _, row in utilization_df.iterrows():
                util_lines.append(f"*{row['Month']}*: {row['Utilization %']} utilization ({row['Numbers Used (1+ SMS)']} of {row['Active Pool']} numbers used)")

        # Current totals (last row)
        latest = total_rows.iloc[-1]
        latest_month = latest['Month']

        # Monthly cost may not be present in total rows; derive from country rows if missing.
        if 'Monthly Cost (EUR)' in total_rows.columns:
            latest_monthly_cost = float(latest.get('Monthly Cost (EUR)', 0))
        else:
            month_rows = breakdown_df[
                (breakdown_df['Month'] == latest_month) &
                (breakdown_df['Country'].isin(COUNTRY_PRICES.keys()))
            ]
            latest_monthly_cost = sum(
                float(r.get('Active Numbers', 0)) * COUNTRY_PRICES.get(r.get('Country'), 0)
                for _, r in month_rows.iterrows()
            )

        message = {
            "blocks": [
                {
                    "type": "header",
                    "text": {
                        "type": "plain_text",
                        "text": "📊 Vonage Weekly Analytics Report",
                        "emoji": True
                    }
                },
                {
                    "type": "section",
                    "text": {
                        "type": "mrkdwn",
                        "text": f"*Report Generated:* {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                    }
                },
                {
                    "type": "divider"
                },
                {
                    "type": "section",
                    "text": {
                        "type": "mrkdwn",
                        "text": f"*📞 Current Active Numbers:* {latest['Active Numbers']}\n*💰 Monthly Cost:* {latest_monthly_cost:.2f} EUR"
                    }
                },
                {
                    "type": "divider"
                },
                {
                    "type": "section",
                    "text": {
                        "type": "mrkdwn",
                        "text": "*Monthly Breakdown:*\n" + "\n".join(summary_lines)
                    }
                },
                {
                    "type": "divider"
                },
                {
                    "type": "section",
                    "text": {
                        "type": "mrkdwn",
                        "text": "*Number Utilization:*\n" + "\n".join(util_lines)
                    }
                },
                {
                    "type": "divider"
                },
                {
                    "type": "context",
                    "elements": [
                        {
                            "type": "mrkdwn",
                            "text": f"📎 Full report saved: `{excel_filename}`"
                        }
                    ]
                }
            ]
        }

        response = requests.post(SLACK_WEBHOOK_URL, json=message)
        if response.status_code == 200:
            print("   Slack notification sent successfully!")
            return True
        else:
            print(f"   Slack notification failed: {response.status_code}")
            return False

    except Exception as e:
        print(f"   Error sending Slack notification: {e}")
        return False

# ==========================================
# PART 1: NUMBERS (Inventory + Audit)
# ==========================================

def get_price(country_code):
    return COUNTRY_PRICES.get(country_code, 0.0)

def format_date_iso(date_str):
    if not date_str or "Unknown" in date_str: return None
    try: return date_str.split('T')[0]
    except: return None

def get_auth_header_audit():
    credentials = f"{API_KEY}:{API_SECRET}"
    encoded = base64.b64encode(credentials.encode()).decode()
    return {"Authorization": f"Basic {encoded}"}

def fetch_numbers_data():
    print("   > Fetching Active Inventory...")
    active_map = {}
    index = 1
    has_more = True
    
    while has_more:
        try:
            url = f"{BASE_URL_REST}/account/numbers"
            params = {"api_key": API_KEY, "api_secret": API_SECRET, "index": index, "size": 100}
            resp = requests.get(url, params=params)
            data = resp.json()
            
            if 'numbers' in data:
                for num in data['numbers']:
                    msisdn = num.get('msisdn')
                    c = num.get('country')
                    active_map[msisdn] = {
                        "Country": c,
                        "MSISDN": msisdn,
                        "Type": num.get('type'),
                        "Features": ", ".join(num.get('features', [])),
                        "Status": "Active",
                        "Purchase Date": None,
                        "Cancel Date": None,
                        "Price (EUR)": get_price(c)
                    }
                if index * 100 < data.get('count', 0): index += 1
                else: has_more = False
            else: has_more = False
        except: has_more = False

    print("   > Fetching Audit Logs for Dates...")
    url = f"{BASE_URL_API}/beta/audit/events"
    headers = get_auth_header_audit()
    page = 1
    has_more_audit = True
    
    while has_more_audit:
        params = {"date_from": f"{START_DATE_HISTORY}T00:00:00", "page": page, "size": 100}
        try:
            resp = requests.get(url, headers=headers, params=params)
            if resp.status_code != 200: break
            
            data = resp.json()
            events = data.get('_embedded', {}).get('events', [])
            if not events: break

            for event in events:
                etype = event.get('event_type')
                if etype in ['NUMBER_ASSIGN', 'NUMBER_RELEASE']:
                    ctx = event.get('context', {})
                    msisdn = ctx.get('number') or ctx.get('msisdn')
                    
                    if msisdn:
                        d_occured = format_date_iso(event.get('created_at'))
                        
                        if msisdn not in active_map:
                            active_map[msisdn] = {
                                "Country": ctx.get('country', 'Unknown'),
                                "MSISDN": msisdn,
                                "Type": "Unknown (Cancelled)",
                                "Features": "Unknown",
                                "Status": "Cancelled",
                                "Purchase Date": None,
                                "Cancel Date": None,
                                "Price (EUR)": get_price(ctx.get('country'))
                            }
                        
                        if etype == 'NUMBER_ASSIGN':
                            active_map[msisdn]['Purchase Date'] = d_occured
                        elif etype == 'NUMBER_RELEASE':
                            active_map[msisdn]['Cancel Date'] = d_occured
                            active_map[msisdn]['Status'] = "Cancelled"
            
            if 'next' not in data.get('_links', {}): has_more_audit = False
            else: 
                page += 1
                time.sleep(0.1)
        except: break
        
    df = pd.DataFrame(active_map.values())
    return df

# ==========================================
# PART 2: SMS REPORT
# ==========================================

def fetch_sms_data():
    print("   > Fetching SMS Report (With Body)...")
    end_date_sms = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    
    url = f"{BASE_URL_API}/v2/reports"
    headers = {
        "Authorization": f"Basic {base64.b64encode(f'{API_KEY}:{API_SECRET}'.encode()).decode()}",
        "Content-Type": "application/json"
    }
    
    payload = {
        "account_id": API_KEY, "product": "SMS", "direction": "Inbound",
        "date_start": START_DATE_SMS, "date_end": end_date_sms,
        "include_message": True
    }
    
    resp = requests.post(url, headers=headers, json=payload)
    if resp.status_code not in [200, 202]:
        print(f"❌ SMS Request Failed: {resp.status_code}")
        return None
        
    data = resp.json()
    rid = data.get('request_id') or data.get('report_id')
    print(f"   > Report Requested. ID: {rid}")
    
    status = "PENDING"
    dl_url = None
    start = time.time()
    
    while status not in ['COMPLETED', 'SUCCESS', 'completed', 'success']:
        if time.time() - start > 600: 
            print("❌ Timeout waiting for report.")
            return None
        time.sleep(5)
        
        check = requests.get(f"{BASE_URL_API}/v2/reports/{rid}", headers=headers)
        if check.status_code != 200: continue
        
        d = check.json()
        status = d.get('status') or d.get('request_status')
        print(f"   > Status: {status}...")
        
        if status in ['COMPLETED', 'SUCCESS', 'completed', 'success']:
            dl_url = d.get('_links', {}).get('download_report', {}).get('href')

    if not dl_url: return None
    
    # Download & Robust Decode
    print("   > Downloading File...")
    r_file = requests.get(dl_url, headers=headers)
    file_content = r_file.content
    try: return pd.read_csv(io.BytesIO(file_content), compression='gzip')
    except:
        try: return pd.read_csv(io.BytesIO(file_content), compression='zip')
        except:
            try: return pd.read_csv(io.BytesIO(file_content))
            except:
                try: return pd.read_csv(io.BytesIO(file_content), encoding='latin-1')
                except: return None

# ==========================================
# PART 3: CALCULATION FUNCTIONS
# ==========================================

def build_number_master_list(df_nums):
    """
    Build a clean master list of numbers, distinguishing between reliable audit data
    and 'Pre-tracking' numbers (bought before Nov 1, 2025).

    Logic:
    - No Purchase Date -> 'Pre-tracking' (assumed active since before Nov 2025)
    - Status=Cancelled and No Cancel Date -> 'Cancelled (Date Unknown)'
    """
    df = df_nums.copy()
    
    # Identify Pre-tracking numbers (active or cancelled, but no assignment event in audit log)
    mask_pretracking = df['Purchase Date'].isna()
    df['Is Pre-tracking'] = mask_pretracking
    
    # Identify numbers with unknown cancellation dates
    mask_unknown_cancel = (df['Status'] == 'Cancelled') & (df['Cancel Date'].isna())
    df['Has Known Cancel Date'] = ~mask_unknown_cancel

    # Summary printing
    pre_count = mask_pretracking.sum()
    unknown_cancel_count = mask_unknown_cancel.sum()
    print(f"   > Master List Built:")
    print(f"     - Total Numbers Found: {len(df)}")
    print(f"     - Pre-tracking Numbers (No Purchase Date): {pre_count}")
    print(f"     - Cancelled Numbers (Known Date): {((df['Status'] == 'Cancelled') & df['Cancel Date'].notna()).sum()}")
    print(f"     - Cancelled Numbers (Unknown Date): {unknown_cancel_count}")

    return df


def generate_month_range(start_date, end_date):
    """Generate list of months from start to end date."""
    months = []
    current = pd.Period(start_date, freq='M')
    end = pd.Period(end_date, freq='M')
    while current <= end:
        months.append(current)
        current = current + 1
    return months


def calculate_monthly_breakdown(df_nums):
    """
    Calculate active numbers AT each month using date range logic.
    For each month/country:
    - Active = (purchased <= month_end OR Pre-tracking) AND (status=Active OR cancelled >= month_start)
    - Cancelled this month = cancel_date within month
    - New activations = purchase_date within month
    - Net change = new - cancelled
    - % Share = country_active / total_active * 100
    """
    # Generate all months from START_DATE to now
    start_month = pd.Period(START_DATE_HISTORY, freq='M')
    end_month = pd.Period(datetime.now(), freq='M')
    all_months = generate_month_range(start_month, end_month)

    countries = ['US', 'CA', 'GB']
    rows = []

    for month in all_months:
        month_start = month.start_time
        month_end = month.end_time

        month_totals = {'active': 0, 'cancelled': 0, 'new': 0, 'cost': 0}
        country_data = {}

        for country in countries:
            country_nums = df_nums[df_nums['Country'] == country]

            # Active at this month: (purchased <= month_end OR Is Pre-tracking) AND (still active OR cancelled >= month_start)
            active_mask = (
                ((country_nums['Purchase Date'] <= month_end) | (country_nums['Is Pre-tracking'] == True)) &
                (
                    (country_nums['Status'] == 'Active') |
                    (country_nums['Cancel Date'] >= month_start)
                )
            )
            active_count = active_mask.sum()

            # Cancelled this month: cancel_date within this month (only if date is known)
            cancelled_mask = (
                (country_nums['Cancel Date'] >= month_start) &
                (country_nums['Cancel Date'] <= month_end)
            )
            cancelled_count = cancelled_mask.sum()

            # New activations this month: purchase_date within this month
            new_mask = (
                (country_nums['Purchase Date'] >= month_start) &
                (country_nums['Purchase Date'] <= month_end)
            )
            new_count = new_mask.sum()

            # Net change
            net_change = new_count - cancelled_count

            # Monthly cost (active numbers * price)
            monthly_cost = active_count * get_price(country)

            country_data[country] = {
                'active': active_count,
                'cancelled': cancelled_count,
                'new': new_count,
                'net': net_change,
                'cost': monthly_cost
            }

            month_totals['active'] += active_count
            month_totals['cancelled'] += cancelled_count
            month_totals['new'] += new_count
            month_totals['cost'] += monthly_cost

        # Calculate % share and add rows
        for country in countries:
            data = country_data[country]
            share_pct = (data['active'] / month_totals['active'] * 100) if month_totals['active'] > 0 else 0

            rows.append({
                'Month': str(month),
                'Country': country,
                'Active Numbers': data['active'],
                'New Activations': data['new'],
                'Net Change': data['net'],
                '% Share': f"{share_pct:.1f}%"
            })

        # Add month total row
        rows.append({
            'Month': str(month),
            'Country': '>>> TOTAL',
            'Active Numbers': month_totals['active'],
            'New Activations': month_totals['new'],
            'Net Change': month_totals['new'] - month_totals['cancelled'],
            '% Share': '100.0%'
        })

    return pd.DataFrame(rows)


def calculate_daily_growth(df_nums):
    """
    Calculate daily activations, cancellations, and net change.
    """
    # Get date range
    start_date = pd.to_datetime(START_DATE_HISTORY).date()
    end_date = datetime.now().date()

    all_dates = pd.date_range(start=start_date, end=end_date, freq='D')

    rows = []
    for date in all_dates:
        date_val = date.date()

        # Activations: purchase_date = this day
        activations = ((df_nums['Purchase Date'].dt.date == date_val)).sum()

        # Cancellations: cancel_date = this day
        cancellations = ((df_nums['Cancel Date'].dt.date == date_val)).sum()

        # Net change
        net_change = activations - cancellations

        rows.append({
            'Date': date_val,
            'Activations': activations,
            'Cancellations': cancellations,
            'Net Change': net_change
        })

    return pd.DataFrame(rows)


def calculate_number_utilization(df_nums, df_sms):
    """
    Calculate number utilization: what % of active numbers received at least 1 SMS per month.

    For each month shows:
    - Active Pool: numbers that were active during that month
    - Numbers Used: numbers that received at least 1 SMS
    - Numbers Unused: numbers that received 0 SMS
    - Utilization %: Numbers Used / Active Pool * 100
    """
    # Generate all months from START_DATE to now
    start_month = pd.Period(START_DATE_HISTORY, freq='M')
    end_month = pd.Period(datetime.now(), freq='M')
    all_months = generate_month_range(start_month, end_month)

    # Get set of numbers that received SMS per month
    sms_receivers_by_month = {}
    if 'to' in df_sms.columns and 'Month' in df_sms.columns:
        for m in df_sms['Month'].unique():
            month_sms = df_sms[df_sms['Month'] == m]
            sms_receivers_by_month[m] = set(month_sms['to'].unique())

    rows = []
    for month in all_months:
        month_start = month.start_time
        month_end = month.end_time

        # Active Pool: (Bought <= End OR Is Pre-tracking) AND (Active OR Cancelled >= Start)
        active_nums = df_nums[
            ((df_nums['Purchase Date'] <= month_end) | (df_nums['Is Pre-tracking'] == True)) &
            ((df_nums['Status'] == 'Active') | (df_nums['Cancel Date'] >= month_start))
        ]
        pool_size = len(active_nums)
        active_numbers_set = set(active_nums['MSISDN'].astype(str))

        # Numbers that received at least 1 SMS this month
        receivers_this_month = sms_receivers_by_month.get(month, set())
        # Convert to strings for comparison
        receivers_this_month = set(str(x) for x in receivers_this_month)

        # Find overlap: active numbers that received SMS
        numbers_used = len(active_numbers_set & receivers_this_month)
        numbers_unused = pool_size - numbers_used

        # Utilization %
        utilization_pct = (numbers_used / pool_size * 100) if pool_size > 0 else 0

        rows.append({
            'Month': str(month),
            'Active Pool': pool_size,
            'Numbers Used (1+ SMS)': numbers_used,
            'Numbers Unused (0 SMS)': numbers_unused,
            'Utilization %': f"{utilization_pct:.1f}%"
        })

    return pd.DataFrame(rows)


def calculate_top_senders(df_sms, top_n=10):
    """
    Calculate top N senders per month with rank.
    """
    if 'from' not in df_sms.columns or 'Month' not in df_sms.columns:
        return pd.DataFrame()

    rows = []
    for m in sorted(df_sms['Month'].unique()):
        sub = df_sms[df_sms['Month'] == m]
        top = sub['from'].value_counts().head(top_n).reset_index()
        top.columns = ['Sender', 'Message Count']

        for rank, (_, row) in enumerate(top.iterrows(), 1):
            rows.append({
                'Month': str(m),
                'Rank': rank,
                'Sender': row['Sender'],
                'Message Count': row['Message Count']
            })

    return pd.DataFrame(rows)


def calculate_heavy_users(df_sms, top_n=10):
    """
    Calculate top N receiving phone numbers per month with rank.
    """
    if 'to' not in df_sms.columns or 'Month' not in df_sms.columns:
        return pd.DataFrame()

    rows = []
    for m in sorted(df_sms['Month'].unique()):
        sub = df_sms[df_sms['Month'] == m]
        top = sub['to'].value_counts().head(top_n).reset_index()
        top.columns = ['Phone Number', 'Message Count']

        for rank, (_, row) in enumerate(top.iterrows(), 1):
            rows.append({
                'Month': str(m),
                'Rank': rank,
                'Phone Number': row['Phone Number'],
                'Message Count': row['Message Count']
            })

    return pd.DataFrame(rows)




# ==========================================
# PART 4: EXPORT FUNCTIONS
# ==========================================

def pivot_by_month(df, value_col, label_col):
    """
    Pivot data so months appear as columns side by side.
    Returns a DataFrame with Rank and month columns (Label, Count) for each month.
    """
    if df.empty:
        return pd.DataFrame()

    months = df['Month'].unique()
    max_rank = df['Rank'].max() if 'Rank' in df.columns else 10

    result_data = {'Rank': list(range(1, max_rank + 1))}

    for month in sorted(months):
        month_data = df[df['Month'] == month].sort_values('Rank')
        labels = []
        counts = []
        for rank in range(1, max_rank + 1):
            row = month_data[month_data['Rank'] == rank]
            if not row.empty:
                labels.append(row[label_col].values[0])
                counts.append(row['Message Count'].values[0])
            else:
                labels.append('')
                counts.append('')

        result_data[f'{month} {label_col}'] = labels
        result_data[f'{month} Count'] = counts

    return pd.DataFrame(result_data)


def export_to_excel(filename, breakdown_df, daily_growth_df, number_utilization_df, sms_senders_df, sms_heavy_df, MasterListSummary):
    """Export all metrics to a multi-sheet Excel file with formatting & charts."""
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.chart import LineChart, Reference, BarChart

    try:
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # 1. Summary Sheet (Metadata & Data Quality)
            summary_data = [
                ['Report Generated', datetime.now().strftime('%Y-%m-%d %H:%M')],
                ['Tracking Period Start', START_DATE_HISTORY],
                ['', ''],
                ['DATA COVERAGE SUMMARY', ''],
                ['Total Numbers Tracked', MasterListSummary.get('total', 0)],
                ['Pre-tracking Numbers (Old)', MasterListSummary.get('pre', 0)],
                ['Cancelled (Date Known)', MasterListSummary.get('cancel_known', 0)],
                ['Cancelled (Date Unknown)', MasterListSummary.get('cancel_unknown', 0)],
                ['', ''],
                ['SMS DATA', ''],
                ['Total Inbound SMS', MasterListSummary.get('sms_total', 0)],
            ]
            pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False, header=False)
            ws_summary = writer.sheets['Summary']
            ws_summary.column_dimensions['A'].width = 30
            ws_summary.column_dimensions['B'].width = 30
            ws_summary['A4'].font = Font(bold=True)
            ws_summary['A10'].font = Font(bold=True)

            # 2. Monthly Breakdown with yellow highlighting
            breakdown_df.to_excel(writer, sheet_name='Monthly Breakdown', index=False)
            ws_breakdown = writer.sheets['Monthly Breakdown']
            yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')

            for row_idx, row in enumerate(ws_breakdown.iter_rows(min_row=2, max_row=ws_breakdown.max_row), start=2):
                if row[1].value == '>>> TOTAL':
                    for cell in row: cell.fill = yellow_fill

            # 3. Daily Growth with Chart
            daily_growth_df.to_excel(writer, sheet_name='Daily Growth', index=False)
            ws_daily = writer.sheets['Daily Growth']
            chart1 = LineChart()
            chart1.title = "Daily Activations / Cancellations"
            chart1.style = 10
            chart1.x_axis.title = "Date"
            chart1.y_axis.title = "Count"
            chart1.width = 25
            chart1.height = 10
            data1 = Reference(ws_daily, min_col=2, min_row=1, max_col=3, max_row=len(daily_growth_df) + 1)
            dates1 = Reference(ws_daily, min_col=1, min_row=2, max_row=len(daily_growth_df) + 1)
            chart1.add_data(data1, titles_from_data=True)
            chart1.set_categories(dates1)
            ws_daily.add_chart(chart1, "F2")

            # 4. Number Utilization
            if not number_utilization_df.empty:
                number_utilization_df.to_excel(writer, sheet_name='Number Utilization', index=False)

            # 5. SMS Top Senders
            if not sms_senders_df.empty:
                pivot_by_month(sms_senders_df, 'Message Count', 'Sender').to_excel(writer, sheet_name='SMS Top Senders', index=False)

            # 6. SMS Heavy Users
            if not sms_heavy_df.empty:
                pivot_by_month(sms_heavy_df, 'Message Count', 'Phone Number').to_excel(writer, sheet_name='SMS Heavy Users', index=False)

        print(f"   Excel saved: {filename}")
        return True
    except Exception as e:
        print(f"   Error creating Excel: {e}")
        return False




# ==========================================
# PART 5: ANALYSIS & REPORT GENERATION
# ==========================================

def run_analysis_and_export(df_nums, df_sms):
    """Main analysis and export function following the 7-step plan."""
    # --- PREPARE DATA ---
    df_nums['Purchase Date'] = pd.to_datetime(df_nums['Purchase Date'], errors='coerce')
    df_nums['Cancel Date'] = pd.to_datetime(df_nums['Cancel Date'], errors='coerce')

    # STEP 4: Build number master list (honest about unknowns)
    print(f"\n[STEP 4] Building Master List...")
    df_nums = build_number_master_list(df_nums)
    
    # Prepare summary for Excel
    MasterListSummary = {
        'total': len(df_nums),
        'pre': df_nums['Is Pre-tracking'].sum(),
        'cancel_known': ((df_nums['Status'] == 'Cancelled') & df_nums['Cancel Date'].notna()).sum(),
        'cancel_unknown': ((df_nums['Status'] == 'Cancelled') & df_nums['Cancel Date'].isna()).sum(),
        'sms_total': len(df_sms)
    }

    date_col = next((c for c in df_sms.columns if 'date' in c.lower() or 'timestamp' in c.lower()), None)
    if date_col:
        df_sms['dt'] = pd.to_datetime(df_sms[date_col])
        df_sms['Month'] = df_sms['dt'].dt.to_period('M')
    else:
        df_sms['Month'] = "Unknown"

    print("\n" + "="*60)
    print("       VONAGE ANALYTICS REPORT (v2 - Clean Build)")
    print("="*60)

    # ---------------------------------------------------------
    # STEP 5: MONTHLY METRICS (Inventory, Cost, Churn)
    # ---------------------------------------------------------
    print(f"\n[STEP 5] Calculating Monthly Inventory & Cost Metrics")
    print("-" * 60)
    breakdown_df = calculate_monthly_breakdown(df_nums)
    print(breakdown_df.to_string(index=False))

    print(f"\n\n[STEP 5.1] Calculating Daily Growth (Activations/Cancellations)")
    print("-" * 60)
    daily_growth_df = calculate_daily_growth(df_nums)
    recent = daily_growth_df.tail(14)
    print(recent.to_string(index=False))
    print(f"\n   (Showing last 14 days. Full data in export files.)")

    # ---------------------------------------------------------
    # STEP 6: SMS METRICS (Utilization, Top Senders, Heavy Users)
    # ---------------------------------------------------------
    print(f"\n\n[STEP 6] Calculating Monthly SMS Metrics")
    print("-" * 60)
    
    number_utilization_df = calculate_number_utilization(df_nums, df_sms)
    if not number_utilization_df.empty:
        print("\n   > Number Utilization:")
        print(number_utilization_df.to_string(index=False))

    sms_senders_df = calculate_top_senders(df_sms, top_n=10)
    if not sms_senders_df.empty:
        print("\n   > Top Senders (Top 10 per Month):")
        # Just show summary for first month to save space in console
        print(sms_senders_df.head(10).to_string(index=False))
        print("   ...")

    sms_heavy_df = calculate_heavy_users(df_sms, top_n=10)
    if not sms_heavy_df.empty:
        print("\n   > Heavy Users (Top 10 Receivers per Month):")
        print(sms_heavy_df.head(10).to_string(index=False))
        print("   ...")

    # ---------------------------------------------------------
    # STEP 7: EXPORT & NOTIFY
    # ---------------------------------------------------------
    print("\n" + "="*60)
    print("       [STEP 7] GENERATING EXPORT FILES")
    print("="*60)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_filename = os.path.join(SCRIPT_DIR, f"vonage_analytics_{timestamp}")
    excel_filename = f"{base_filename}.xlsx"

    # Export Excel (CSV exports removed per user request)
    export_to_excel(excel_filename, breakdown_df, daily_growth_df,
                    number_utilization_df, sms_senders_df, sms_heavy_df, MasterListSummary)

    # Send Slack notification
    print("\n   Sending Slack notification...")
    send_slack_notification(breakdown_df, number_utilization_df, excel_filename)

    print(f"\n   All exports complete.")

# ==========================================
# SNAPSHOT MANAGEMENT
# ==========================================

def load_snapshot():
    """Load existing snapshot if available."""
    if os.path.exists(SNAPSHOT_FILE):
        df = pd.read_csv(SNAPSHOT_FILE)
        print(f"   Snapshot loaded: {len(df)} records from {SNAPSHOT_FILE}")
        return df
    print("   No snapshot found. First run - will create one.")
    return None


def merge_with_snapshot(df_fresh, df_snapshot):
    """
    Merge fresh API data with snapshot to keep historical data stable.

    Rules:
    - Numbers already in snapshot: keep snapshot data (stable history)
    - Numbers in API but not snapshot: add them (new numbers)
    - Numbers in snapshot but not API and Status=Active: update to Cancelled
      (number was removed since last run)
    - Cancel dates from snapshot are preserved (never overwritten)
    """
    if df_snapshot is None:
        print(f"   First run - saving {len(df_fresh)} records as initial snapshot.")
        df_fresh.to_csv(SNAPSHOT_FILE, index=False)
        return df_fresh

    snapshot_msisdns = set(df_snapshot['MSISDN'].astype(str))
    fresh_msisdns = set(df_fresh['MSISDN'].astype(str))

    # New numbers: in API but not in snapshot
    new_msisdns = fresh_msisdns - snapshot_msisdns
    new_numbers = df_fresh[df_fresh['MSISDN'].astype(str).isin(new_msisdns)]
    print(f"   New numbers since last snapshot: {len(new_numbers)}")

    # Numbers that disappeared from active inventory: were Active in snapshot,
    # now not in API active list -> mark as Cancelled with today's date
    fresh_active_msisdns = set(df_fresh[df_fresh['Status'] == 'Active']['MSISDN'].astype(str))
    snapshot_active = df_snapshot[df_snapshot['Status'] == 'Active'].copy()
    disappeared = snapshot_active[~snapshot_active['MSISDN'].astype(str).isin(fresh_active_msisdns)]

    if len(disappeared) > 0:
        today = datetime.now().strftime('%Y-%m-%d')
        df_snapshot.loc[disappeared.index, 'Status'] = 'Cancelled'
        # Only set cancel date if not already set
        no_cancel_date = df_snapshot.loc[disappeared.index, 'Cancel Date'].isna()
        df_snapshot.loc[disappeared.index[no_cancel_date], 'Cancel Date'] = today
        print(f"   Numbers cancelled since last run: {len(disappeared)} (cancel date: {today})")

    # Update status of snapshot numbers that are still active in API
    # (in case they were previously marked cancelled by inference but are actually still active)
    still_active = df_snapshot['MSISDN'].astype(str).isin(fresh_active_msisdns)
    df_snapshot.loc[still_active, 'Status'] = 'Active'
    df_snapshot.loc[still_active, 'Cancel Date'] = None

    # Merge: snapshot + new numbers
    df_merged = pd.concat([df_snapshot, new_numbers], ignore_index=True)

    # Save updated snapshot
    df_merged.to_csv(SNAPSHOT_FILE, index=False)
    print(f"   Snapshot updated: {len(df_merged)} total records (saved to {SNAPSHOT_FILE})")

    return df_merged


# ==========================================
# MAIN EXECUTION
# ==========================================
if __name__ == "__main__":
    print("\n--- VONAGE ANALYTICS PIPELINE STARTED ---")
    validate_config()

    # 1. Fetch fresh data from API
    print(f"\n[STEP 1] Fetching data from Vonage API...")
    df_fresh = fetch_numbers_data()
    # Save raw data in script dir
    df_fresh.to_csv(os.path.join(SCRIPT_DIR, "vonage_numbers_raw.csv"), index=False)
    print(f"   Fresh API data: {len(df_fresh)} records")

    # 2. Merge with snapshot for stable historical data
    print("\n[STEP 2] Merging with snapshot...")
    df_snapshot = load_snapshot()
    df_numbers = merge_with_snapshot(df_fresh, df_snapshot)

    # 3. Fetch SMS data
    print("\n[STEP 3] Fetching SMS data...")
    df_sms = fetch_sms_data()
    if df_sms is not None:
        df_sms.to_csv(os.path.join(SCRIPT_DIR, "vonage_sms_raw.csv"), index=False)
        print(f"   SMS data: {len(df_sms)} records")
    else:
        print("   SMS data: None (API error or no data)")

    # 4. Analyze & Export
    print("\n[STEP 4] Running analysis...")
    if not df_numbers.empty:
        # Create empty SMS dataframe if no SMS data
        if df_sms is None or df_sms.empty:
            print("   Warning: No SMS data available. Running numbers analysis only.")
            df_sms = pd.DataFrame(columns=['from', 'to', 'date_received', 'Month'])
        run_analysis_and_export(df_numbers, df_sms)
    else:
        print("   No numbers data available.")

    print("\n--- PIPELINE FINISHED ---")
